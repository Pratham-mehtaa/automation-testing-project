"""
generate_bug_report.py — Parse test results and generate an Excel bug report.
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def create_bug_report():
    """Generate an Excel bug report from the test results."""

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════
    # Sheet 1: Bug Report
    # ═══════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Bug Report"

    # Styling
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    high_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    medium_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    low_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    severity_font_white = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    severity_font_black = Font(name="Calibri", bold=True, size=10, color="000000")
    body_font = Font(name="Calibri", size=10)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    center_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        "Bug ID", "Test Case ID", "Module", "User Type", "Severity",
        "Bug Title", "Steps to Reproduce", "Expected Result",
        "Actual Result", "Status", "Screenshot"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # Bug data (extracted from test results)
    bugs = [
        {
            "bug_id": "BUG-001",
            "test_id": "TC-PRD-012",
            "module": "Product/Inventory",
            "user": "problem_user",
            "severity": "High",
            "title": "All product images show the same 404 error image",
            "steps": "1. Login as problem_user\n2. Navigate to inventory page\n3. Observe product images",
            "expected": "Each product should display its own unique product image",
            "actual": "All 6 products show the same image: sl-404.168b1cce10384b857a6f.jpg (a 404/broken image placeholder)",
            "status": "Open",
            "screenshot": "FAIL_test_problem_user_images_broken.png",
        },
        {
            "bug_id": "BUG-002",
            "test_id": "TC-PRD-014",
            "module": "Product/Inventory",
            "user": "problem_user",
            "severity": "High",
            "title": "Sort Z-A does not work — products remain in A-Z order",
            "steps": "1. Login as problem_user\n2. Select 'Name (Z to A)' from sort dropdown\n3. Observe product order",
            "expected": "Products should be sorted in reverse alphabetical order (Z to A)",
            "actual": "Products remain in A-Z order: ['Sauce Labs Backpack', 'Sauce Labs Bike Light', ...] instead of reversed",
            "status": "Open",
            "screenshot": "FAIL_test_problem_user_sort_name_za.png",
        },
        {
            "bug_id": "BUG-003",
            "test_id": "TC-PRD-015",
            "module": "Product/Inventory",
            "user": "problem_user",
            "severity": "High",
            "title": "Sort Price (Low to High) does not sort correctly",
            "steps": "1. Login as problem_user\n2. Select 'Price (low to high)' from sort dropdown\n3. Observe product prices",
            "expected": "Prices should be in ascending order: $7.99, $9.99, $15.99, $15.99, $29.99, $49.99",
            "actual": "Prices remain unsorted: $29.99, $9.99, $15.99, $49.99, $7.99, $15.99",
            "status": "Open",
            "screenshot": "FAIL_test_problem_user_sort_price_lohi.png",
        },
        {
            "bug_id": "BUG-004",
            "test_id": "TC-PRD-016",
            "module": "Product/Inventory",
            "user": "problem_user",
            "severity": "High",
            "title": "Sort Price (High to Low) does not sort correctly",
            "steps": "1. Login as problem_user\n2. Select 'Price (high to low)' from sort dropdown\n3. Observe product prices",
            "expected": "Prices should be in descending order: $49.99, $29.99, $15.99, $15.99, $9.99, $7.99",
            "actual": "Prices remain unsorted: $29.99, $9.99, $15.99, $49.99, $7.99, $15.99",
            "status": "Open",
            "screenshot": "FAIL_test_problem_user_sort_price_hilo.png",
        },
        {
            "bug_id": "BUG-005",
            "test_id": "TC-CHK-009",
            "module": "Checkout",
            "user": "problem_user",
            "severity": "Critical",
            "title": "Checkout form first name field only retains last character typed",
            "steps": "1. Login as problem_user\n2. Add item to cart\n3. Go to checkout\n4. Type 'John' in First Name field\n5. Check the field value",
            "expected": "First Name field should contain 'John'",
            "actual": "First Name field contains only 'e' (the last character of the previous field interaction). Input is corrupted.",
            "status": "Open",
            "screenshot": "FAIL_test_problem_user_checkout_info.png",
        },
        {
            "bug_id": "BUG-006",
            "test_id": "TC-CHK-010",
            "module": "Checkout",
            "user": "problem_user",
            "severity": "Critical",
            "title": "Checkout fails with 'Last Name is required' despite entering data",
            "steps": "1. Login as problem_user\n2. Add item to cart\n3. Go to checkout\n4. Fill in first name, last name, postal code\n5. Click Continue",
            "expected": "Checkout should proceed to step 2 (Overview)",
            "actual": "Error displayed: 'Error: Last Name is required' — the last name field does not accept input correctly",
            "status": "Open",
            "screenshot": "FAIL_test_problem_user_complete_checkout.png",
        },
        {
            "bug_id": "BUG-007",
            "test_id": "TC-CHK-011",
            "module": "Checkout",
            "user": "error_user",
            "severity": "Critical",
            "title": "Last Name field does not accept input — remains empty",
            "steps": "1. Login as error_user\n2. Add item to cart\n3. Go to checkout\n4. Type 'Doe' in Last Name field\n5. Check the field value",
            "expected": "Last Name field should contain 'Doe'",
            "actual": "Last Name field is empty ('') — the typed value is not retained",
            "status": "Open",
            "screenshot": "FAIL_test_error_user_checkout_info.png",
        },
        {
            "bug_id": "BUG-008",
            "test_id": "TC-CHK-012",
            "module": "Checkout",
            "user": "error_user",
            "severity": "Critical",
            "title": "Checkout cannot be completed — Finish button not clickable",
            "steps": "1. Login as error_user\n2. Add item to cart\n3. Complete checkout info\n4. Click Finish on overview page",
            "expected": "Order should be confirmed with success message",
            "actual": "TimeoutException — Finish button cannot be found/clicked. Checkout flow is broken.",
            "status": "Open",
            "screenshot": "FAIL_test_error_user_complete_checkout.png",
        },
        {
            "bug_id": "BUG-009",
            "test_id": "TC-PRD-019",
            "module": "Product/Inventory",
            "user": "error_user",
            "severity": "High",
            "title": "Sorting triggers JavaScript alert: 'Sorting is broken!'",
            "steps": "1. Login as error_user\n2. Select 'Name (Z to A)' from sort dropdown\n3. Observe behavior",
            "expected": "Products should be sorted Z to A",
            "actual": "JavaScript alert appears: 'Sorting is broken! This error has been reported to Backtrace.' — All sort options (Z-A, Low-High, High-Low) trigger this alert.",
            "status": "Open",
            "screenshot": "FAIL_test_error_user_sort_name_za.png",
        },
        {
            "bug_id": "BUG-010",
            "test_id": "TC-PRD-020",
            "module": "Product/Inventory",
            "user": "error_user",
            "severity": "High",
            "title": "Sort Price Low-to-High triggers JS error alert",
            "steps": "1. Login as error_user\n2. Select 'Price (low to high)' sort option",
            "expected": "Products sorted by price ascending",
            "actual": "Alert: 'Sorting is broken! This error has been reported to Backtrace.'",
            "status": "Open",
            "screenshot": "FAIL_test_error_user_sort_price_lohi.png",
        },
        {
            "bug_id": "BUG-011",
            "test_id": "TC-PRD-021",
            "module": "Product/Inventory",
            "user": "error_user",
            "severity": "High",
            "title": "Sort Price High-to-Low triggers JS error alert",
            "steps": "1. Login as error_user\n2. Select 'Price (high to low)' sort option",
            "expected": "Products sorted by price descending",
            "actual": "Alert: 'Sorting is broken! This error has been reported to Backtrace.'",
            "status": "Open",
            "screenshot": "FAIL_test_error_user_sort_price_hilo.png",
        },
    ]

    # Write bug data
    for row_idx, bug in enumerate(bugs, 2):
        values = [
            bug["bug_id"], bug["test_id"], bug["module"], bug["user"],
            bug["severity"], bug["title"], bug["steps"], bug["expected"],
            bug["actual"], bug["status"], bug["screenshot"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = wrap_alignment
            cell.border = thin_border

        # Color severity cell
        sev_cell = ws.cell(row=row_idx, column=5)
        sev_cell.alignment = center_alignment
        if bug["severity"] == "Critical":
            sev_cell.fill = critical_fill
            sev_cell.font = severity_font_white
        elif bug["severity"] == "High":
            sev_cell.fill = high_fill
            sev_cell.font = severity_font_white
        elif bug["severity"] == "Medium":
            sev_cell.fill = medium_fill
            sev_cell.font = severity_font_black
        elif bug["severity"] == "Low":
            sev_cell.fill = low_fill
            sev_cell.font = severity_font_black

        # Center Bug ID, Test ID, Module, User, Status
        for c in [1, 2, 3, 4, 10]:
            ws.cell(row=row_idx, column=c).alignment = center_alignment

    # Column widths
    col_widths = {
        1: 10, 2: 14, 3: 18, 4: 18, 5: 12,
        6: 50, 7: 45, 8: 40, 9: 50, 10: 10, 11: 45,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[chr(64 + col)].width = width

    # ═══════════════════════════════════════════════════════════════
    # Sheet 2: Test Summary
    # ═══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Test Summary")

    summary_headers = ["Module", "Total Tests", "Passed", "Failed", "Pass Rate"]
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for col, header in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    summary_data = [
        ("Login", 10, 9, 1, "90%"),
        ("Product/Inventory", 21, 14, 7, "67%"),
        ("Cart", 11, 11, 0, "100%"),
        ("Checkout", 12, 3, 9, "25%"),
        ("TOTAL", 54, 36, 18, "67%"),
    ]

    for row_idx, (module, total, passed, failed, rate) in enumerate(summary_data, 2):
        values = [module, total, passed, failed, rate]
        for col_idx, value in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10, bold=(row_idx == len(summary_data) + 1))
            cell.alignment = center_alignment
            cell.border = thin_border

        # Highlight pass/fail
        ws2.cell(row=row_idx, column=3).fill = pass_fill
        if failed > 0:
            ws2.cell(row=row_idx, column=4).fill = fail_fill

    # Bold the TOTAL row
    total_row = len(summary_data) + 1
    for col in range(1, 6):
        ws2.cell(row=total_row, column=col).font = Font(name="Calibri", size=10, bold=True)

    for col in range(1, 6):
        ws2.column_dimensions[chr(64 + col)].width = 18

    # ═══════════════════════════════════════════════════════════════
    # Sheet 3: Bugs by User Type
    # ═══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Bugs by User Type")

    user_headers = ["User Type", "Bugs Found", "Critical", "High", "Medium", "Low", "Affected Areas"]
    for col, header in enumerate(user_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    user_data = [
        ("standard_user", 0, 0, 0, 0, 0, "None — all tests pass"),
        ("locked_out_user", 0, 0, 0, 0, 0, "Expected behavior — locked out"),
        ("problem_user", 6, 2, 4, 0, 0, "Images, Sorting (all options), Checkout form fields"),
        ("performance_glitch_user", 0, 0, 0, 0, 0, "Slow login (~5s) but functional"),
        ("error_user", 5, 2, 3, 0, 0, "Sorting (JS alert), Checkout (last name, finish button)"),
        ("visual_user", 0, 0, 0, 0, 0, "Not fully tested (visual bugs need pixel comparison)"),
    ]

    for row_idx, data in enumerate(user_data, 2):
        for col_idx, value in enumerate(data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = center_alignment if col_idx < 7 else wrap_alignment
            cell.border = thin_border

    for col in range(1, 8):
        ws3.column_dimensions[chr(64 + col)].width = 20
    ws3.column_dimensions["G"].width = 50

    # Save
    output_path = os.path.join(
        os.path.dirname(__file__), "reports", "SauceDemo_Bug_Report.xlsx"
    )
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"✅ Bug report saved: {output_path}")
    return output_path


if __name__ == "__main__":
    create_bug_report()
